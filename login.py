#!/usr/bin/env python3
"""
Fluxo HTTP para Clinica Agil.

O script:
- autentica sem Selenium
- exporta o relatorio de vendas para o periodo informado
- exporta a planilha de pacientes
- persiste pacientes em SQLite com carga incremental por reexecucao
"""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import os
import re
import sqlite3
import sys
import unicodedata
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import requests

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


BASE_URL = "https://app2.clinicaagil.com.br"


def _runtime_root() -> Path:
    override = os.getenv("MIRELLA_RUNTIME_ROOT")
    if override:
        return Path(override).resolve()
    return Path(__file__).resolve().parent


RUNTIME_ROOT = _runtime_root()
DEFAULT_ENV_FILE = RUNTIME_ROOT / ".env"
DEFAULT_TIMEOUT = 60
DEFAULT_EMAIL = "consultores@agregarnegocios.com.br"
DEFAULT_SENHA = "@Agregar12"
DEFAULT_DATA_VENDAS = "01/01/2024"
DEFAULT_DATA_PACIENTES = "01/01/2024"
DEFAULT_OUTPUT_DIR = RUNTIME_ROOT / "exports"
DEFAULT_DB_PATH = RUNTIME_ROOT / "mirella_pacientes.sqlite3"

LOGIN_CONFIG: Dict[str, Any] = {
    "base_url": BASE_URL,
    "endpoints": {
        "login": "/login",
        "relatorios": "/financeiro/relatorio/index",
        "vendas": "/financeiro/relatorio/vendas",
        "pacientes": "/pacientes/exportar_xls",
    },
    "credenciais": {
        "email": DEFAULT_EMAIL,
        "senha": DEFAULT_SENHA,
    },
}

PATIENT_HEADER_TO_COLUMN = {
    "id": "patient_id",
    "nome": "nome",
    "data_nasc": "data_nasc",
    "telefone_1": "telefone_1",
    "telefone_2": "telefone_2",
    "telefone_3": "telefone_3",
    "matricula": "matricula",
    "convenio": "convenio",
    "sexo": "sexo",
    "etnia": "etnia",
    "responsaveis": "responsaveis",
    "nome_da_mae": "nome_mae",
    "cpf": "cpf",
    "identidade": "identidade",
    "cep": "cep",
    "endereco": "endereco",
    "e_mail": "email",
    "profissao": "profissao",
    "status": "status",
    "cidade": "cidade",
    "bairro": "bairro",
    "plano": "plano",
    "cpf_responsavel": "cpf_responsavel",
    "cns": "cns",
}

PATIENT_DATA_COLUMNS = [
    "nome",
    "data_nasc",
    "telefone_1",
    "telefone_2",
    "telefone_3",
    "matricula",
    "convenio",
    "sexo",
    "etnia",
    "responsaveis",
    "nome_mae",
    "cpf",
    "identidade",
    "cep",
    "endereco",
    "email",
    "profissao",
    "status",
    "cidade",
    "bairro",
    "plano",
    "cpf_responsavel",
    "cns",
]

PLACEHOLDER_STRINGS = {"'", "-", "--", "+55", "+55 "}


def _load_env_file(env_path: Path) -> None:
    if not env_path.exists():
        return
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("export "):
            line = line[len("export "):].strip()
        if "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip()
        if not key:
            continue
        if len(value) >= 2 and value[0] == value[-1] and value[0] in ("'", '"'):
            value = value[1:-1]
        if key not in os.environ:
            os.environ[key] = value


def _env_int(name: str, default: int) -> int:
    raw = os.getenv(name)
    if raw is None or not raw.strip():
        return default
    try:
        return int(raw.strip())
    except ValueError:
        return default


def _hoje_br() -> str:
    return datetime.now().strftime("%d/%m/%Y")


def _validar_data_br(valor: str) -> str:
    try:
        data = datetime.strptime(valor.strip(), "%d/%m/%Y")
    except ValueError as exc:
        raise ValueError(f"Data invalida '{valor}'. Use o formato dd/mm/aaaa.") from exc
    return data.strftime("%d/%m/%Y")


def _data_br_para_iso(valor: str) -> str:
    return datetime.strptime(_validar_data_br(valor), "%d/%m/%Y").strftime("%Y-%m-%d")


def _validar_periodo(data_de: str, data_ate: str, contexto: str) -> Tuple[str, str]:
    inicio = datetime.strptime(_validar_data_br(data_de), "%d/%m/%Y")
    fim = datetime.strptime(_validar_data_br(data_ate), "%d/%m/%Y")
    if inicio > fim:
        raise ValueError(f"Periodo invalido para {contexto}: data inicial maior que data final.")
    return inicio.strftime("%d/%m/%Y"), fim.strftime("%d/%m/%Y")


def _normalizar_header(valor: Any) -> str:
    texto = "" if valor is None else str(valor).strip().lower()
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    texto = re.sub(r"[^a-z0-9]+", "_", texto)
    return texto.strip("_")


def _normalizar_celula(valor: Any) -> Optional[str]:
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y")
    texto = str(valor).strip()
    return texto or None


def _prompt_data(rotulo: str, valor_padrao: str, habilitado: bool) -> str:
    if not habilitado:
        return _validar_data_br(valor_padrao)
    while True:
        resposta = input(f"{rotulo} [{valor_padrao}]: ").strip()
        valor = resposta or valor_padrao
        try:
            return _validar_data_br(valor)
        except ValueError as exc:
            print(exc)


def _garantir_resposta_excel(resp: requests.Response, endpoint: str) -> None:
    if resp.status_code != 200:
        raise RuntimeError(f"Falha em {endpoint}: HTTP {resp.status_code}")
    content_type = (resp.headers.get("Content-Type") or "").lower()
    if "excel" in content_type or resp.content[:2] == b"PK":
        return
    raise RuntimeError(
        f"Resposta inesperada em {endpoint}. "
        f"Content-Type: {resp.headers.get('Content-Type', '<vazio>')}"
    )


def _celula_para_numero(valor: Any) -> float:
    if valor is None or valor == "":
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    if isinstance(valor, str):
        texto = valor.strip().replace("R$", "").strip()
        texto = texto.replace(".", "").replace(",", ".")
        try:
            return float(texto)
        except ValueError:
            return 0.0
    return 0.0


def _inserir_coluna_diferenca_vendas(caminho_xlsx: Path) -> None:
    if load_workbook is None:
        raise RuntimeError("openpyxl nao instalado. Instale com: pip install openpyxl")
    workbook = load_workbook(str(caminho_xlsx), read_only=False)
    worksheet = workbook.active
    worksheet.insert_cols(9)
    worksheet.cell(row=3, column=9).value = "Descontos"
    for row_idx in range(4, worksheet.max_row + 1):
        valor_g = _celula_para_numero(worksheet.cell(row=row_idx, column=7).value)
        valor_h = _celula_para_numero(worksheet.cell(row=row_idx, column=8).value)
        worksheet.cell(row=row_idx, column=9).value = valor_g - valor_h
    workbook.save(str(caminho_xlsx))
    workbook.close()


def _extrair_pacientes_do_xlsx(conteudo: bytes) -> List[Dict[str, Optional[str]]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl nao instalado. Instale com: pip install openpyxl")

    workbook = load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
    worksheet = workbook.active

    cabecalhos_norm: Optional[List[str]] = None
    registros: List[Dict[str, Optional[str]]] = []

    for row in worksheet.iter_rows(values_only=True):
        if not any(value not in (None, "") for value in row):
            continue

        if cabecalhos_norm is None:
            cabecalhos_norm = [_normalizar_header(value) for value in row]
            continue

        registro: Dict[str, Optional[str]] = {}
        for indice, header_norm in enumerate(cabecalhos_norm):
            if header_norm not in PATIENT_HEADER_TO_COLUMN:
                continue
            coluna = PATIENT_HEADER_TO_COLUMN[header_norm]
            registro[coluna] = _normalizar_celula(row[indice] if indice < len(row) else None)

        patient_id = registro.get("patient_id")
        if not patient_id:
            continue

        for coluna in PATIENT_DATA_COLUMNS:
            registro.setdefault(coluna, None)

        registros.append(registro)

    workbook.close()
    return registros


def _texto_ou_none(valor: Any) -> Optional[str]:
    texto = _normalizar_celula(valor)
    if texto is None:
        return None
    if texto in PLACEHOLDER_STRINGS:
        return None
    return texto


def _normalizar_documento(valor: Any, tamanho: Optional[int] = None) -> Optional[str]:
    texto = _texto_ou_none(valor)
    if not texto:
        return None
    digits = re.sub(r"\D", "", texto)
    if not digits:
        return None
    if tamanho is not None and len(digits) != tamanho:
        return None
    if set(digits) == {"0"}:
        return None
    return digits


def _normalizar_email(valor: Any) -> Optional[str]:
    texto = _texto_ou_none(valor)
    if not texto:
        return None
    return texto.lower()


def _normalizar_telefone(valor: Any) -> Optional[str]:
    texto = _texto_ou_none(valor)
    if not texto:
        return None
    digits = re.sub(r"\D", "", texto)
    if not digits:
        return None
    significant = digits[2:] if digits.startswith("55") else digits
    if len(significant) < 10:
        return None
    if set(significant) == {"0"}:
        return None
    return digits


def _normalizar_data_br_ou_none(valor: Any) -> Optional[str]:
    texto = _texto_ou_none(valor)
    if not texto:
        return None
    try:
        return datetime.strptime(texto, "%d/%m/%Y").strftime("%Y-%m-%d")
    except ValueError:
        return None


def _curar_registro_paciente(row: sqlite3.Row) -> Dict[str, Any]:
    patient_id = int(str(row["patient_id"]).strip())
    sexo = _texto_ou_none(row["sexo"])
    if sexo:
        sexo = sexo.upper()

    patient = {
        "patient_id": patient_id,
        "nome": _texto_ou_none(row["nome"]) or f"Paciente {patient_id}",
        "data_nascimento": _normalizar_data_br_ou_none(row["data_nasc"]),
        "sexo": sexo,
        "cpf": _normalizar_documento(row["cpf"], tamanho=11),
        "identidade": _texto_ou_none(row["identidade"]),
        "status": _texto_ou_none(row["status"]),
        "convenio": _texto_ou_none(row["convenio"]),
        "plano": _texto_ou_none(row["plano"]),
        "profissao": _texto_ou_none(row["profissao"]),
        "responsavel_nome": _texto_ou_none(row["responsaveis"]),
        "responsavel_cpf": _normalizar_documento(row["cpf_responsavel"], tamanho=11),
        "nome_mae": _texto_ou_none(row["nome_mae"]),
        "cns": _normalizar_documento(row["cns"]),
        "row_hash": row["row_hash"],
        "imported_at": row["imported_at"],
        "first_seen_at": row["first_seen_at"],
        "last_seen_at": row["last_seen_at"],
        "source_period_start": row["source_period_start"],
        "source_period_end": row["source_period_end"],
        "source_file_name": row["source_file_name"],
    }

    contacts: List[Dict[str, Any]] = []
    phone_contacts: List[Dict[str, Any]] = []
    for label, raw_value in (
        ("telefone_1", row["telefone_1"]),
        ("telefone_2", row["telefone_2"]),
        ("telefone_3", row["telefone_3"]),
    ):
        contact_value = _texto_ou_none(raw_value)
        contact_norm = _normalizar_telefone(raw_value)
        if contact_value and contact_norm:
            phone_contacts.append(
                {
                    "patient_id": patient_id,
                    "contact_type": "phone",
                    "contact_label": label,
                    "contact_value": contact_value,
                    "contact_value_norm": contact_norm,
                    "is_primary": 0,
                    "imported_at": row["imported_at"],
                }
            )

    if phone_contacts:
        phone_contacts[0]["is_primary"] = 1
        contacts.extend(phone_contacts)

    email_value = _texto_ou_none(row["email"])
    email_norm = _normalizar_email(row["email"])
    if email_value and email_norm:
        contacts.append(
            {
                "patient_id": patient_id,
                "contact_type": "email",
                "contact_label": "email_principal",
                "contact_value": email_value,
                "contact_value_norm": email_norm,
                "is_primary": 1,
                "imported_at": row["imported_at"],
            }
        )

    address = {
        "patient_id": patient_id,
        "cep": _normalizar_documento(row["cep"], tamanho=8),
        "endereco": _texto_ou_none(row["endereco"]),
        "bairro": _texto_ou_none(row["bairro"]),
        "cidade": _texto_ou_none(row["cidade"]),
        "imported_at": row["imported_at"],
    }

    return {
        "patient": patient,
        "contacts": contacts,
        "address": address,
    }


class SQLitePatientStore:
    def __init__(self, db_path: Path, logger: logging.Logger) -> None:
        self.db_path = Path(db_path)
        self.logger = logger
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self.conn = sqlite3.connect(str(self.db_path))
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA foreign_keys = ON")
        self._criar_schema()

    def close(self) -> None:
        self.conn.close()

    def _criar_schema(self) -> None:
        colunas_dados = ",\n                ".join(f"{coluna} TEXT" for coluna in PATIENT_DATA_COLUMNS)
        self.conn.executescript(
            f"""
            CREATE TABLE IF NOT EXISTS sync_state (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS patients_latest (
                patient_id TEXT PRIMARY KEY,
                row_hash TEXT NOT NULL,
                imported_at TEXT NOT NULL,
                first_seen_at TEXT NOT NULL,
                last_seen_at TEXT NOT NULL,
                source_period_start TEXT NOT NULL,
                source_period_end TEXT NOT NULL,
                source_file_name TEXT,
                raw_payload_json TEXT NOT NULL,
                {colunas_dados}
            );

            CREATE TABLE IF NOT EXISTS patient_versions (
                patient_id TEXT NOT NULL,
                row_hash TEXT NOT NULL,
                imported_at TEXT NOT NULL,
                source_period_start TEXT NOT NULL,
                source_period_end TEXT NOT NULL,
                source_file_name TEXT,
                raw_payload_json TEXT NOT NULL,
                PRIMARY KEY (patient_id, row_hash)
            );

            CREATE INDEX IF NOT EXISTS idx_patient_versions_imported_at
            ON patient_versions(imported_at);

            CREATE TABLE IF NOT EXISTS patient_import_runs (
                run_id INTEGER PRIMARY KEY AUTOINCREMENT,
                imported_at TEXT NOT NULL,
                source_period_start TEXT NOT NULL,
                source_period_end TEXT NOT NULL,
                source_file_name TEXT,
                total_rows INTEGER NOT NULL,
                inserted_rows INTEGER NOT NULL,
                updated_rows INTEGER NOT NULL,
                unchanged_rows INTEGER NOT NULL,
                versions_added INTEGER NOT NULL
            );

            CREATE TABLE IF NOT EXISTS patients (
                patient_id INTEGER PRIMARY KEY,
                nome TEXT NOT NULL,
                data_nascimento TEXT,
                sexo TEXT,
                cpf TEXT,
                identidade TEXT,
                status TEXT,
                convenio TEXT,
                plano TEXT,
                profissao TEXT,
                responsavel_nome TEXT,
                responsavel_cpf TEXT,
                nome_mae TEXT,
                cns TEXT,
                row_hash TEXT NOT NULL,
                imported_at TEXT NOT NULL,
                first_seen_at TEXT NOT NULL,
                last_seen_at TEXT NOT NULL,
                source_period_start TEXT NOT NULL,
                source_period_end TEXT NOT NULL,
                source_file_name TEXT
            );

            CREATE INDEX IF NOT EXISTS idx_patients_cpf ON patients(cpf);
            CREATE INDEX IF NOT EXISTS idx_patients_status ON patients(status);
            CREATE INDEX IF NOT EXISTS idx_patients_convenio ON patients(convenio);

            CREATE TABLE IF NOT EXISTS patient_contacts (
                contact_id INTEGER PRIMARY KEY AUTOINCREMENT,
                patient_id INTEGER NOT NULL,
                contact_type TEXT NOT NULL,
                contact_label TEXT NOT NULL,
                contact_value TEXT NOT NULL,
                contact_value_norm TEXT NOT NULL,
                is_primary INTEGER NOT NULL DEFAULT 0,
                imported_at TEXT NOT NULL,
                UNIQUE(patient_id, contact_type, contact_label),
                UNIQUE(patient_id, contact_type, contact_value_norm),
                FOREIGN KEY(patient_id) REFERENCES patients(patient_id) ON DELETE CASCADE
            );

            CREATE INDEX IF NOT EXISTS idx_patient_contacts_patient ON patient_contacts(patient_id);
            CREATE INDEX IF NOT EXISTS idx_patient_contacts_lookup
            ON patient_contacts(contact_type, contact_value_norm);

            CREATE TABLE IF NOT EXISTS patient_addresses (
                patient_id INTEGER PRIMARY KEY,
                cep TEXT,
                endereco TEXT,
                bairro TEXT,
                cidade TEXT,
                imported_at TEXT NOT NULL,
                FOREIGN KEY(patient_id) REFERENCES patients(patient_id) ON DELETE CASCADE
            );

            CREATE INDEX IF NOT EXISTS idx_patient_addresses_city ON patient_addresses(cidade);
            CREATE INDEX IF NOT EXISTS idx_patient_addresses_neighborhood ON patient_addresses(bairro);

            CREATE VIEW IF NOT EXISTS vw_patients_complete AS
            SELECT
                p.patient_id,
                p.nome,
                p.data_nascimento,
                p.sexo,
                p.cpf,
                p.status,
                p.convenio,
                p.plano,
                p.profissao,
                a.cep,
                a.endereco,
                a.bairro,
                a.cidade,
                phone.contact_value AS telefone_principal,
                email.contact_value AS email_principal,
                p.imported_at,
                p.source_period_start,
                p.source_period_end
            FROM patients p
            LEFT JOIN patient_addresses a
                ON a.patient_id = p.patient_id
            LEFT JOIN patient_contacts phone
                ON phone.patient_id = p.patient_id
               AND phone.contact_type = 'phone'
               AND phone.is_primary = 1
            LEFT JOIN patient_contacts email
                ON email.patient_id = p.patient_id
               AND email.contact_type = 'email'
               AND email.is_primary = 1;
            """
        )
        self.conn.commit()

    def get_state(self, key: str) -> Optional[str]:
        row = self.conn.execute("SELECT value FROM sync_state WHERE key = ?", (key,)).fetchone()
        return None if row is None else row["value"]

    def set_state(self, key: str, value: str) -> None:
        agora = datetime.now().isoformat(timespec="seconds")
        self.conn.execute(
            """
            INSERT INTO sync_state (key, value, updated_at)
            VALUES (?, ?, ?)
            ON CONFLICT(key) DO UPDATE SET
                value = excluded.value,
                updated_at = excluded.updated_at
            """,
            (key, value, agora),
        )

    def get_default_start_date(self) -> Optional[str]:
        return self.get_state("patients_last_sync_end_date")

    def _registrar_import_run(
        self,
        imported_at: str,
        source_period_start_iso: str,
        source_period_end_iso: str,
        source_file_name: str,
        resumo: Dict[str, int],
    ) -> None:
        self.conn.execute(
            """
            INSERT INTO patient_import_runs (
                imported_at,
                source_period_start,
                source_period_end,
                source_file_name,
                total_rows,
                inserted_rows,
                updated_rows,
                unchanged_rows,
                versions_added
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                imported_at,
                source_period_start_iso,
                source_period_end_iso,
                source_file_name,
                resumo["total"],
                resumo["inserted"],
                resumo["updated"],
                resumo["unchanged"],
                resumo["versions_added"],
            ),
        )

    def rebuild_curated_tables(self) -> Dict[str, int]:
        rows = self.conn.execute(
            "SELECT * FROM patients_latest ORDER BY CAST(patient_id AS INTEGER)"
        ).fetchall()

        cursor = self.conn.cursor()
        cursor.execute("DELETE FROM patient_contacts")
        cursor.execute("DELETE FROM patient_addresses")
        cursor.execute("DELETE FROM patients")

        resumo = {
            "curated_patients": 0,
            "curated_contacts": 0,
            "curated_addresses": 0,
        }

        for row in rows:
            registro = _curar_registro_paciente(row)
            patient = registro["patient"]
            cursor.execute(
                """
                INSERT INTO patients (
                    patient_id,
                    nome,
                    data_nascimento,
                    sexo,
                    cpf,
                    identidade,
                    status,
                    convenio,
                    plano,
                    profissao,
                    responsavel_nome,
                    responsavel_cpf,
                    nome_mae,
                    cns,
                    row_hash,
                    imported_at,
                    first_seen_at,
                    last_seen_at,
                    source_period_start,
                    source_period_end,
                    source_file_name
                ) VALUES (
                    :patient_id,
                    :nome,
                    :data_nascimento,
                    :sexo,
                    :cpf,
                    :identidade,
                    :status,
                    :convenio,
                    :plano,
                    :profissao,
                    :responsavel_nome,
                    :responsavel_cpf,
                    :nome_mae,
                    :cns,
                    :row_hash,
                    :imported_at,
                    :first_seen_at,
                    :last_seen_at,
                    :source_period_start,
                    :source_period_end,
                    :source_file_name
                )
                """,
                patient,
            )
            resumo["curated_patients"] += 1

            address = registro["address"]
            if any(address[coluna] is not None for coluna in ("cep", "endereco", "bairro", "cidade")):
                cursor.execute(
                    """
                    INSERT INTO patient_addresses (
                        patient_id,
                        cep,
                        endereco,
                        bairro,
                        cidade,
                        imported_at
                    ) VALUES (
                        :patient_id,
                        :cep,
                        :endereco,
                        :bairro,
                        :cidade,
                        :imported_at
                    )
                    """,
                    address,
                )
                resumo["curated_addresses"] += 1

            for contact in registro["contacts"]:
                result = cursor.execute(
                    """
                    INSERT OR IGNORE INTO patient_contacts (
                        patient_id,
                        contact_type,
                        contact_label,
                        contact_value,
                        contact_value_norm,
                        is_primary,
                        imported_at
                    ) VALUES (
                        :patient_id,
                        :contact_type,
                        :contact_label,
                        :contact_value,
                        :contact_value_norm,
                        :is_primary,
                        :imported_at
                    )
                    """,
                    contact,
                )
                if result.rowcount:
                    resumo["curated_contacts"] += 1

        return resumo

    def upsert_patients(
        self,
        rows: Sequence[Dict[str, Optional[str]]],
        source_period_start: str,
        source_period_end: str,
        source_file_name: str,
    ) -> Dict[str, int]:
        imported_at = datetime.now().isoformat(timespec="seconds")
        source_period_start_iso = _data_br_para_iso(source_period_start)
        source_period_end_iso = _data_br_para_iso(source_period_end)

        insert_sql = f"""
            INSERT INTO patients_latest (
                patient_id,
                row_hash,
                imported_at,
                first_seen_at,
                last_seen_at,
                source_period_start,
                source_period_end,
                source_file_name,
                raw_payload_json,
                {", ".join(PATIENT_DATA_COLUMNS)}
            ) VALUES (
                :patient_id,
                :row_hash,
                :imported_at,
                :first_seen_at,
                :last_seen_at,
                :source_period_start,
                :source_period_end,
                :source_file_name,
                :raw_payload_json,
                {", ".join(f":{coluna}" for coluna in PATIENT_DATA_COLUMNS)}
            )
            ON CONFLICT(patient_id) DO UPDATE SET
                row_hash = excluded.row_hash,
                imported_at = excluded.imported_at,
                first_seen_at = patients_latest.first_seen_at,
                last_seen_at = excluded.last_seen_at,
                source_period_start = excluded.source_period_start,
                source_period_end = excluded.source_period_end,
                source_file_name = excluded.source_file_name,
                raw_payload_json = excluded.raw_payload_json,
                {", ".join(f"{coluna} = excluded.{coluna}" for coluna in PATIENT_DATA_COLUMNS)}
        """

        resumo = {
            "total": 0,
            "inserted": 0,
            "updated": 0,
            "unchanged": 0,
            "versions_added": 0,
        }

        cursor = self.conn.cursor()
        for row in rows:
            patient_id = row["patient_id"]
            payload_json = json.dumps(row, ensure_ascii=False, sort_keys=True)
            row_hash = hashlib.sha256(payload_json.encode("utf-8")).hexdigest()

            existente = cursor.execute(
                "SELECT row_hash FROM patients_latest WHERE patient_id = ?",
                (patient_id,),
            ).fetchone()

            if existente is None:
                resumo["inserted"] += 1
            elif existente["row_hash"] == row_hash:
                resumo["unchanged"] += 1
            else:
                resumo["updated"] += 1

            version_result = cursor.execute(
                """
                INSERT OR IGNORE INTO patient_versions (
                    patient_id,
                    row_hash,
                    imported_at,
                    source_period_start,
                    source_period_end,
                    source_file_name,
                    raw_payload_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    patient_id,
                    row_hash,
                    imported_at,
                    source_period_start_iso,
                    source_period_end_iso,
                    source_file_name,
                    payload_json,
                ),
            )
            if version_result.rowcount:
                resumo["versions_added"] += 1

            payload: Dict[str, Optional[str]] = {
                "patient_id": patient_id,
                "row_hash": row_hash,
                "imported_at": imported_at,
                "first_seen_at": imported_at,
                "last_seen_at": imported_at,
                "source_period_start": source_period_start_iso,
                "source_period_end": source_period_end_iso,
                "source_file_name": source_file_name,
                "raw_payload_json": payload_json,
            }
            for coluna in PATIENT_DATA_COLUMNS:
                payload[coluna] = row.get(coluna)

            cursor.execute(insert_sql, payload)
            resumo["total"] += 1

        self.set_state("patients_last_sync_end_date", source_period_end)
        self.set_state("patients_last_sync_start_date", source_period_start)
        self.set_state("patients_last_sync_at", imported_at)
        self.set_state("patients_last_sync_rows", str(resumo["total"]))
        self._registrar_import_run(
            imported_at=imported_at,
            source_period_start_iso=source_period_start_iso,
            source_period_end_iso=source_period_end_iso,
            source_file_name=source_file_name,
            resumo=resumo,
        )
        resumo.update(self.rebuild_curated_tables())
        self.set_state("patients_curated_count", str(resumo["curated_patients"]))
        self.set_state("patients_curated_contacts", str(resumo["curated_contacts"]))
        self.set_state("patients_curated_addresses", str(resumo["curated_addresses"]))
        self.conn.commit()
        return resumo


class ClinicaAgilHTTPExporter:
    def __init__(
        self,
        email: str,
        senha: str,
        timeout: int,
        output_dir: Path,
        logger: logging.Logger,
    ) -> None:
        self.email = email
        self.senha = senha
        self.timeout = timeout
        self.output_dir = Path(output_dir)
        self.logger = logger
        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/145.0.0.0 Safari/537.36"
                )
            }
        )

    def autenticar(self, cookie_env: Optional[str] = None) -> None:
        if cookie_env:
            self.logger.info("Autenticacao: usando cookie de sessao informado.")
            self.session.headers["Cookie"] = cookie_env
            resposta = self.session.get(
                f"{BASE_URL}{LOGIN_CONFIG['endpoints']['relatorios']}",
                timeout=self.timeout,
            )
            if resposta.status_code == 200:
                return
            raise RuntimeError(f"Cookie de sessao invalido. HTTP {resposta.status_code}")

        self.logger.info("Autenticacao: fazendo login por HTTP.")
        resposta_login = self.session.post(
            f"{BASE_URL}{LOGIN_CONFIG['endpoints']['login']}",
            data={"identity": self.email, "password": self.senha},
            timeout=self.timeout,
        )
        if resposta_login.status_code not in (200, 302):
            raise RuntimeError(f"Falha no login. HTTP {resposta_login.status_code}")

        resposta_relatorios = self.session.get(
            f"{BASE_URL}{LOGIN_CONFIG['endpoints']['relatorios']}",
            timeout=self.timeout,
        )
        if resposta_relatorios.status_code != 200:
            raise RuntimeError(
                "Login retornou sem acesso a tela de relatorios. "
                f"HTTP {resposta_relatorios.status_code}"
            )
        self.logger.info("Autenticacao concluida com sucesso.")

    def _post_excel(self, endpoint: str, payload: Dict[str, str], nome: str) -> bytes:
        self.logger.info("Solicitando '%s' via HTTP...", nome)
        resposta = self.session.post(f"{BASE_URL}{endpoint}", data=payload, timeout=self.timeout)
        _garantir_resposta_excel(resposta, endpoint)
        self.logger.info("Arquivo '%s' recebido (%.1f KB).", nome, len(resposta.content) / 1024.0)
        return resposta.content

    def exportar_vendas(self, data_de: str, data_ate: str) -> Dict[str, Path]:
        conteudo = self._post_excel(
            LOGIN_CONFIG["endpoints"]["vendas"],
            {
                "data_de": data_de,
                "data_ate": data_ate,
                "tipo": "xls",
            },
            "relatorio de vendas",
        )

        tag = f"{_data_br_para_iso(data_de)}_{_data_br_para_iso(data_ate)}"
        destino_dir = self.output_dir / "vendas"
        destino_dir.mkdir(parents=True, exist_ok=True)

        arquivo_profissionais = destino_dir / f"relatorio_vendas_profissionais_{tag}.xlsx"
        arquivo_geral = destino_dir / f"relatorio_vendas_geral_{tag}.xlsx"

        arquivo_profissionais.write_bytes(conteudo)
        arquivo_geral.write_bytes(conteudo)
        _inserir_coluna_diferenca_vendas(arquivo_geral)

        self.logger.info("Vendas salvas em: %s", arquivo_profissionais)
        self.logger.info("Vendas tratadas salvas em: %s", arquivo_geral)
        return {
            "profissionais": arquivo_profissionais,
            "geral": arquivo_geral,
        }

    def exportar_pacientes(self, data_de: str, data_ate: str) -> Tuple[bytes, Path]:
        conteudo = self._post_excel(
            LOGIN_CONFIG["endpoints"]["pacientes"],
            {
                "data_de": data_de,
                "data_ate": data_ate,
            },
            "pacientes",
        )

        destino_dir = self.output_dir / "pacientes"
        destino_dir.mkdir(parents=True, exist_ok=True)
        arquivo_latest = destino_dir / "pacientes_list_latest.xlsx"
        arquivo_latest.write_bytes(conteudo)
        self.logger.info("Planilha mais recente de pacientes salva em: %s", arquivo_latest)
        return conteudo, arquivo_latest


def _resolver_periodo_vendas(args: argparse.Namespace) -> Optional[Tuple[str, str]]:
    if args.somente == "pacientes":
        return None

    pode_perguntar = not args.sem_input and sys.stdin.isatty()
    padrao_de = args.data_vendas_de or os.getenv("MIRELLA_DATA_VENDAS", DEFAULT_DATA_VENDAS)
    padrao_ate = args.data_vendas_ate or os.getenv("MIRELLA_DATA_VENDAS_ATE", _hoje_br())

    data_de = _prompt_data("Data inicial do relatorio de vendas", padrao_de, pode_perguntar)
    data_ate = _prompt_data("Data final do relatorio de vendas", padrao_ate, pode_perguntar)
    return _validar_periodo(data_de, data_ate, "vendas")


def _resolver_periodo_pacientes(args: argparse.Namespace, store: SQLitePatientStore) -> Optional[Tuple[str, str]]:
    if args.somente == "vendas":
        return None

    if args.reprocessar_pacientes:
        padrao_de = (
            args.data_pacientes_de
            or os.getenv("MIRELLA_DATA_PACIENTES", DEFAULT_DATA_PACIENTES)
        )
    else:
        padrao_de = (
            args.data_pacientes_de
            or store.get_default_start_date()
            or os.getenv("MIRELLA_DATA_PACIENTES", DEFAULT_DATA_PACIENTES)
        )

    padrao_ate = args.data_pacientes_ate or os.getenv("MIRELLA_DATA_PACIENTES_ATE", _hoje_br())
    return _validar_periodo(padrao_de, padrao_ate, "pacientes")


def _criar_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Exporta vendas e pacientes da Clinica Agil via HTTP e sincroniza pacientes em SQLite."
    )
    parser.add_argument("--email", default=os.getenv("MIRELLA_EMAIL", DEFAULT_EMAIL))
    parser.add_argument("--senha", default=os.getenv("MIRELLA_SENHA", DEFAULT_SENHA))
    parser.add_argument("--cookie", default=os.getenv("MIRELLA_COOKIE"))
    parser.add_argument("--timeout", type=int, default=_env_int("MIRELLA_TIMEOUT", DEFAULT_TIMEOUT))
    parser.add_argument("--output-dir", default=str(Path(os.getenv("MIRELLA_OUTPUT_DIR", DEFAULT_OUTPUT_DIR))))
    parser.add_argument("--db-path", default=str(Path(os.getenv("MIRELLA_DB_PATH", DEFAULT_DB_PATH))))
    parser.add_argument("--somente", choices=("ambos", "vendas", "pacientes"), default="ambos")
    parser.add_argument("--sem-input", action="store_true", help="Nao pergunta datas no terminal.")
    parser.add_argument("--reprocessar-pacientes", action="store_true", help="Ignora a data da ultima sincronizacao.")
    parser.add_argument("--rebuild-curated", action="store_true", help="Reconstrui as tabelas curadas a partir de patients_latest sem chamar a API.")
    parser.add_argument("--data-vendas-de")
    parser.add_argument("--data-vendas-ate")
    parser.add_argument("--data-pacientes-de")
    parser.add_argument("--data-pacientes-ate")
    return parser


def main() -> None:
    _load_env_file(DEFAULT_ENV_FILE)
    parser = _criar_parser()
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(message)s", datefmt="%H:%M:%S")
    logger = logging.getLogger(__name__)

    logger.info("============================================================")
    logger.info("Exportacao HTTP Clinica Agil")
    logger.info("Ambiente: %s", DEFAULT_ENV_FILE)
    logger.info("Saida: %s", args.output_dir)
    logger.info("SQLite: %s", args.db_path)
    logger.info("Modo: %s", args.somente)
    logger.info("============================================================")

    store = SQLitePatientStore(Path(args.db_path), logger)
    try:
        if args.rebuild_curated:
            resumo_curado = store.rebuild_curated_tables()
            store.set_state("patients_curated_count", str(resumo_curado["curated_patients"]))
            store.set_state("patients_curated_contacts", str(resumo_curado["curated_contacts"]))
            store.set_state("patients_curated_addresses", str(resumo_curado["curated_addresses"]))
            store.conn.commit()
            logger.info(
                "Tabelas curadas reconstruidas: pacientes=%s | contatos=%s | enderecos=%s",
                resumo_curado["curated_patients"],
                resumo_curado["curated_contacts"],
                resumo_curado["curated_addresses"],
            )
            return

        periodo_vendas = _resolver_periodo_vendas(args)
        periodo_pacientes = _resolver_periodo_pacientes(args, store)

        if periodo_vendas:
            logger.info("Periodo de vendas: %s ate %s", periodo_vendas[0], periodo_vendas[1])
        if periodo_pacientes:
            logger.info(
                "Periodo de pacientes: %s ate %s",
                periodo_pacientes[0],
                periodo_pacientes[1],
            )

        exporter = ClinicaAgilHTTPExporter(
            email=args.email,
            senha=args.senha,
            timeout=args.timeout,
            output_dir=Path(args.output_dir),
            logger=logger,
        )

        exporter.autenticar(cookie_env=args.cookie)

        if periodo_vendas:
            arquivos_vendas = exporter.exportar_vendas(*periodo_vendas)
            logger.info("Arquivos de vendas gerados: %s | %s", arquivos_vendas["profissionais"], arquivos_vendas["geral"])

        if periodo_pacientes:
            conteudo_pacientes, arquivo_pacientes = exporter.exportar_pacientes(*periodo_pacientes)
            rows = _extrair_pacientes_do_xlsx(conteudo_pacientes)
            resumo = store.upsert_patients(
                rows=rows,
                source_period_start=periodo_pacientes[0],
                source_period_end=periodo_pacientes[1],
                source_file_name=arquivo_pacientes.name,
            )
            logger.info(
                "Pacientes sincronizados em SQLite: total=%s | novos=%s | atualizados=%s | sem_alteracao=%s | versoes_novas=%s | curated_patients=%s | curated_contacts=%s | curated_addresses=%s",
                resumo["total"],
                resumo["inserted"],
                resumo["updated"],
                resumo["unchanged"],
                resumo["versions_added"],
                resumo["curated_patients"],
                resumo["curated_contacts"],
                resumo["curated_addresses"],
            )

    finally:
        store.close()


def setup_driver(*_args: Any, **_kwargs: Any) -> None:
    raise RuntimeError("login.py agora e um fluxo HTTP. Use a execucao direta deste script.")


def realizar_login(*_args: Any, **_kwargs: Any) -> None:
    raise RuntimeError("login.py agora e um fluxo HTTP. Use a execucao direta deste script.")


def extrair_relatorio_vendas(*_args: Any, **_kwargs: Any) -> None:
    raise RuntimeError("login.py agora e um fluxo HTTP. Use a execucao direta deste script.")


def extrair_pacientes(*_args: Any, **_kwargs: Any) -> None:
    raise RuntimeError("login.py agora e um fluxo HTTP. Use a execucao direta deste script.")


if __name__ == "__main__":
    main()
